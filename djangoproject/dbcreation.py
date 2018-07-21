import django, os, sys

os.environ['DJANGO_SETTINGS_MODULE'] = "onlineproject.settings"
django.setup()
from onlineapp.models import *

import MySQLdb
import click
import openpyxl
from bs4 import BeautifulSoup


@click.group()
def request():
    "Commands for db import/export"
    pass


conn = MySQLdb.connect("localhost", "root", "Aj@y1234")
cursor = conn.cursor()


@request.command()
def cleardata():
    pass


@request.command()
@click.argument('dbname')
# @click.argument('table_spec')

def createdb(dbname):
    try:
        cursor.execute("CREATE DATABASE " + dbname + ";")
        cursor.execute("use " + dbname + ";")

        cursor.execute(
            "CREATE TABLE Students(Name varchar(255),College varchar(255),EmailId varchar(255),DbName varchar(255) PRIMARY KEY);")
        cursor.execute(
            "CREATE TABLE Marks(DbName varchar(255),PRIMARY KEY(DbName),FOREIGN KEY(DbName) REFERENCES STUDENTS(DbName),Transform INT,From_custon_base_26 INT,get_pig_latin INT,top_chars INT,Total INT);")
    except Exception:
        print("database already exists")


@request.command()
@click.argument('dbname')
def dropdb(dbname):
    # cursor.execute("DROP DATABASE IF EXISTS {0}".format(dbname))
    try:
        cursor.execute("DROP DATABASE " + dbname + ";")
    except Exception:
        print("database doesn't exits")


@request.command()
@click.argument('students_file')
@click.argument('marks_file')
def populatedb(students_file, marks_file):
    stud_sheet = openpyxl.load_workbook(students_file)["Current"]
    clg_sheet = openpyxl.load_workbook(students_file)["Colleges"]
    # marks_sheet=openpyxl.load_workbook(marks_file)["Sheet"]

    cursor.execute("use tutdb")
    print("start")
    for row in range(2, clg_sheet.max_row + 1):
        c = College(name=clg_sheet.cell(row, 1).value, acronym=clg_sheet.cell(row, 2).value,
                    location=clg_sheet.cell(row, 3).value, contact=clg_sheet.cell(row, 4).value)
        c.save()

    print("college Inserted table")
    for row in range(2, stud_sheet.max_row + 1):
        c = Student(name=stud_sheet.cell(row, 1).value,
                    college=College.objects.get(acronym=stud_sheet.cell(row, 2).value),
                    email=stud_sheet.cell(row, 3).value, db_folder=stud_sheet.cell(row, 4).value)
        c.save()

    with open(marks_file, "r") as f:
        soup = BeautifulSoup(f, 'html.parser')
    table = soup.find("table")

    data = []
    rows = table.findAll("tr")
    c = rows[0]
    cols = c.find_all('th')
    cols = [x.text.strip() for x in cols]
    data.append(cols[1:])

    for r in rows:
        cols = r.find_all('td')
        cols = [x.text.strip() for x in cols]
        data.append(cols[1:])

    if [] in data:
        data.remove([])

    data.remove(data[0])
    for i in range(len(data)):
        db_name = data[i][0].split("_")[2].lower()
        try:
            col = Student.objects.get(db_folder=db_name)
            mock = MockTestMarks(problem1=int(data[i][1]), problem2=int(data[i][2]), problem3=int(data[i][3]),
                                 problem4=int(data[i][4]), total=int(data[i][5]),
                                 student=Student.objects.get(db_folder=db_name))
            mock.save()
        except Exception as e:
            print(e)


request()
conn.close()
